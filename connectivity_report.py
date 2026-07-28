import os
import sys
from datetime import datetime
sys.path.append(r"C:\Program Files\DIgSILENT\PowerFactory 2024\Python\3.11")
import powerfactory as pf
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# 1. Connect to PowerFactory
# ---------------------------------------------------------------------------
def connect_to_powerfactory():
    app = pf.GetApplication()
    if app is None:
        raise RuntimeError(
            "Could not connect to PowerFactory. Make sure PowerFactory is "
            "installed/licensed and that the Python API path is correctly "
            "added to sys.path (see comments at the top of this script)."
        )
    app.ClearOutputWindow()
    app.PrintPlain("Connected to PowerFactory successfully.")
    return app


# ---------------------------------------------------------------------------
# 2. List projects and study cases
# ---------------------------------------------------------------------------
def list_projects(app):
    user = app.GetCurrentUser()
    projects = user.GetContents("*.IntPrj", 1)  # recursive search

    if not projects:
        print("No projects found under the current user.")
        return []

    print("\nAvailable Projects:")
    for i, prj in enumerate(projects):
        print(f"  [{i}] {prj.loc_name}")
    return projects


def list_study_cases(project):
    """
    Recursively search inside the project for study case objects
    (IntCase), regardless of which folder they sit in.
    """
    study_cases = project.GetContents("*.IntCase", 1)

    if not study_cases:
        print(f"No study cases found in project '{project.loc_name}'.")
        return []

    print(f"\nStudy Cases in project '{project.loc_name}':")
    for i, sc in enumerate(study_cases):
        print(f"  [{i}] {sc.loc_name}")
    return study_cases


def choose_from_list(items, prompt):
    while True:
        choice = input(prompt).strip()
        if choice.isdigit() and 0 <= int(choice) < len(items):
            return items[int(choice)]
        print(f"Invalid selection. Enter a number between 0 and {len(items) - 1}.")


# ---------------------------------------------------------------------------
# 4. Connectivity report generation
# ---------------------------------------------------------------------------
def get_side_label(element, cubicle):
    """
    Return a human-readable label describing which side/terminal of
    'element' this cubicle corresponds to (e.g. HV Side, LV Side,
    Terminal i, Terminal j, etc.).
    """
    cls = element.GetClassName()

    try:
        if cls == "ElmTr2":  # 2-winding transformer
            if cubicle == element.GetCubicle(0):
                return "HV Side"
            if cubicle == element.GetCubicle(1):
                return "LV Side"

        elif cls == "ElmTr3":  # 3-winding transformer
            if cubicle == element.GetCubicle(0):
                return "HV Side"
            if cubicle == element.GetCubicle(1):
                return "MV Side"
            if cubicle == element.GetCubicle(2):
                return "LV Side"

        elif cls == "ElmLne":  # line / cable
            if cubicle == element.GetCubicle(0):
                return "Terminal i (From)"
            if cubicle == element.GetCubicle(1):
                return "Terminal j (To)"

        elif cls == "ElmCoup":  # breaker / coupling between busbars
            if cubicle == element.GetCubicle(0):
                return "Side 1"
            if cubicle == element.GetCubicle(1):
                return "Side 2"

        elif cls in ("ElmSym", "ElmGenstat", "ElmPvsys", "ElmAsm"):
            return "Generator Terminal"

        elif cls == "ElmLod":
            return "Load Terminal"

        elif cls == "ElmShnt":
            return "Shunt/Capacitor Terminal"

    except Exception:
        pass

    return "Connected"


def generate_connectivity_report(app):
    """
    Walk every terminal (bus) in the active study case and, for each
    cubicle on that terminal, record which element is connected and
    on which side.
    """
    terminals = app.GetCalcRelevantObjects("*.ElmTerm")
    records = []

    for term in terminals:
        bus_name = term.loc_name

        substation = ""
        try:
            parent_substation = term.cpSubstat
            substation = parent_substation.loc_name if parent_substation else ""
        except Exception:
            pass

        try:
            un = term.uknom
        except Exception:
            un = None

        cubicles = term.GetContents("*.StaCubic")
        for cub in cubicles:
            connected_elm = cub.obj_id
            if connected_elm is None:
                continue  # empty/unused cubicle

            elm_class = connected_elm.GetClassName()
            elm_name = connected_elm.loc_name
            cub_name = cub.loc_name
            side_label = get_side_label(connected_elm, cub)

            records.append({
                "Bus (Terminal)": bus_name,
                "Substation": substation,
                "Voltage (kV)": un,
                "Cubicle": cub_name,
                "Connected Element": elm_name,
                "Element Type": elm_class,
                "Connection Side": side_label,
            })

    df = pd.DataFrame(records)
    return df


def generate_element_summary(df):
    """
    Pivot the detailed connectivity table so each element (e.g. each
    transformer or line) appears on a single row, with one column per
    side showing exactly which bus + cubicle it connects to.

    Example result for a transformer:
        Element | Type   | HV Side              | LV Side
        T1      | ElmTr2 | Bus1 (Cubicle: Cub_1)| Bus2 (Cubicle: Cub_2)
    """
    if df.empty:
        return pd.DataFrame()

    summary_rows = []
    grouped = df.groupby(["Connected Element", "Element Type"])
    for (elm_name, elm_type), group in grouped:
        row = {"Element": elm_name, "Type": elm_type}
        for _, r in group.iterrows():
            side = r["Connection Side"]
            row[side] = f"{r['Bus (Terminal)']} (Cubicle: {r['Cubicle']})"
        summary_rows.append(row)

    return pd.DataFrame(summary_rows)


# ---------------------------------------------------------------------------
# 5. Export to Excel with basic formatting
# ---------------------------------------------------------------------------
def export_report(df_detail, df_summary, filename=None):
    if filename is None:
        filename = f"Network_Connectivity_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df_detail.to_excel(writer, sheet_name="Detailed Connectivity", index=False)
        df_summary.to_excel(writer, sheet_name="Element Summary", index=False)

    wb = load_workbook(filename)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

        ws.freeze_panes = "A2"

    wb.save(filename)
    print(f"\nReport saved to: {os.path.abspath(filename)}")
    return filename


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    app = connect_to_powerfactory()

    # --- Select and activate project ---
    projects = list_projects(app)
    if not projects:
        return
    project = choose_from_list(projects, "\nEnter the index of the project to activate: ")

    # Activate project first so its study cases are properly resolvable
    # Note: PowerFactory's Activate() returns 0 = success, 1 = already active
    # (not a real error), and other non-zero codes for genuine failures.
    err = project.Activate()
    if err == 1:
        print(f"\nProject '{project.loc_name}' was already active.")
    elif err:
        raise RuntimeError(f"Failed to activate project '{project.loc_name}' (error code {err}).")
    else:
        print(f"\nProject '{project.loc_name}' activated.")

    # --- Select and activate study case ---
    study_cases = list_study_cases(project)
    if not study_cases:
        return
    study_case = choose_from_list(study_cases, "\nEnter the index of the study case to activate: ")

    err = study_case.Activate()
    if err == 1:
        print(f"Study case '{study_case.loc_name}' was already active.")
    elif err:
        raise RuntimeError(f"Failed to activate study case '{study_case.loc_name}' (error code {err}).")
    else:
        print(f"Study case '{study_case.loc_name}' activated.")

    # --- Generate report ---
    print("\nGenerating Network Connectivity Report...")
    df_detail = generate_connectivity_report(app)

    if df_detail.empty:
        print("No connectivity data found. Check that the study case contains a valid network model.")
        return

    df_summary = generate_element_summary(df_detail)

    print(
        f"\nFound {len(df_detail)} connection records across "
        f"{df_detail['Bus (Terminal)'].nunique()} buses and "
        f"{df_summary.shape[0]} network elements."
    )

    export_report(df_detail, df_summary)


if __name__ == "__main__":
    main()